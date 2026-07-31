from __future__ import annotations

import csv
import io
import json
import re
import zipfile
from collections.abc import Callable
from pathlib import Path
from typing import Any

from .models import public_result_dicts

RESULT_EXPORT_FIELDS = [
    "scan_id",
    "host",
    "port",
    "protocol",
    "state",
    "service_name",
    "banner",
    "evidence",
    "evidence_files",
    "error",
    "tags",
    "note",
    "created_at",
]


def format_results_json(job: dict[str, Any], results: list[dict[str, Any]]) -> str:
    return json.dumps({"job": job, "results": public_result_dicts(results)}, indent=2)


def format_results_csv(results: list[dict[str, Any]]) -> str:
    output = io.StringIO()
    writer = csv.DictWriter(output, fieldnames=RESULT_EXPORT_FIELDS, extrasaction="ignore", lineterminator="\n")
    writer.writeheader()
    for result in public_result_dicts(results):
        row = dict(result)
        row["tags"] = json.dumps(row.get("tags") or [], ensure_ascii=False)
        row["evidence_files"] = json.dumps(row.get("evidence_files") or [], ensure_ascii=False)
        writer.writerow(row)
    return output.getvalue()


def format_results_csv_bundle(
    job: dict[str, Any],
    results: list[dict[str, Any]],
    *,
    load_evidence: Callable[[str], tuple[dict[str, Any], Path] | None],
) -> bytes:
    bundle_results = public_result_dicts(results)
    assets: list[tuple[str, Path]] = []
    manifest_evidence: list[dict[str, Any]] = []
    seen: set[str] = set()

    for result in bundle_results:
        bundled_evidence: list[dict[str, Any]] = []
        for evidence in result.get("evidence_files") or []:
            item = dict(evidence)
            evidence_id = str(item.get("id") or "")
            if evidence_id and evidence_id not in seen:
                loaded = load_evidence(evidence_id)
                if loaded:
                    metadata, path = loaded
                    safe_id = re.sub(r"[^A-Za-z0-9_.-]+", "_", evidence_id)
                    extension = _extension_for_mime_type(str(metadata.get("mime_type") or ""))
                    archive_path = f"evidence/{safe_id}{extension}"
                    item["bundle_path"] = archive_path
                    assets.append((archive_path, path))
                    manifest_evidence.append(item)
                    seen.add(evidence_id)
            elif evidence_id in seen:
                existing = next(
                    (entry for entry in manifest_evidence if entry.get("id") == evidence_id),
                    None,
                )
                if existing and existing.get("bundle_path"):
                    item["bundle_path"] = existing["bundle_path"]
            bundled_evidence.append(item)
        result["evidence_files"] = bundled_evidence

    output = io.BytesIO()
    with zipfile.ZipFile(output, mode="w", compression=zipfile.ZIP_DEFLATED) as archive:
        archive.writestr("results.csv", format_results_csv(bundle_results).encode("utf-8-sig"))
        archive.writestr(
            "manifest.json",
            json.dumps(
                {"job": job, "evidence_files": manifest_evidence},
                ensure_ascii=False,
                indent=2,
            ).encode("utf-8"),
        )
        for archive_path, source_path in assets:
            archive.write(source_path, archive_path)
    return output.getvalue()


def format_results_xlsx(
    job: dict[str, Any],
    results: list[dict[str, Any]],
    *,
    load_evidence: Callable[[str], tuple[dict[str, Any], Path] | None],
) -> bytes:
    try:
        from openpyxl import Workbook
        from openpyxl.drawing.image import Image as ExcelImage
        from openpyxl.styles import Alignment, Font, PatternFill
        from PIL import Image as PillowImage
        from PIL import ImageOps
    except ImportError as exc:
        raise RuntimeError("Excel export requires openpyxl and Pillow; install with: pip install -e .") from exc

    workbook = Workbook()
    workbook.properties.title = f"Scaprobe Scan {job.get('id', '')}".strip()
    workbook.properties.creator = "Scaprobe"
    results_sheet = workbook.active
    results_sheet.title = "Results"
    result_headers = [
        "Scan ID",
        "Host",
        "Port",
        "Protocol",
        "State",
        "Service",
        "Banner",
        "Evidence",
        "Tags",
        "Note",
        "Created At",
        "Image Count",
    ]
    results_sheet.append(result_headers)
    _style_excel_header(results_sheet, Font, PatternFill, Alignment)

    public_results = public_result_dicts(results)
    for result in public_results:
        evidence_files = result.get("evidence_files") or []
        results_sheet.append(
            [
                _excel_text(result.get("scan_id")),
                _excel_text(result.get("host")),
                result.get("port"),
                _excel_text(result.get("protocol")),
                _excel_text(result.get("state")),
                _excel_text(result.get("service_name")),
                _excel_text(result.get("banner")),
                _excel_text(result.get("evidence")),
                _excel_text(", ".join(result.get("tags") or [])),
                _excel_text(result.get("note")),
                _excel_text(result.get("created_at")),
                len(evidence_files),
            ]
        )

    results_sheet.freeze_panes = "A2"
    results_sheet.auto_filter.ref = results_sheet.dimensions
    results_sheet.column_dimensions["A"].width = 38
    results_sheet.column_dimensions["B"].width = 24
    for column in ("C", "D", "E", "F", "L"):
        results_sheet.column_dimensions[column].width = 14
    for column in ("G", "H", "I", "J"):
        results_sheet.column_dimensions[column].width = 36
    results_sheet.column_dimensions["K"].width = 22
    for row in results_sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    evidence_sheet = workbook.create_sheet("Evidence")
    evidence_headers = [
        "Host",
        "Port",
        "Protocol",
        "Service",
        "Type",
        "File Name",
        "SHA-256",
        "Source URL",
        "Image",
    ]
    evidence_sheet.append(evidence_headers)
    _style_excel_header(evidence_sheet, Font, PatternFill, Alignment)
    evidence_sheet.freeze_panes = "A2"
    evidence_sheet.column_dimensions["A"].width = 24
    evidence_sheet.column_dimensions["B"].width = 12
    evidence_sheet.column_dimensions["C"].width = 12
    evidence_sheet.column_dimensions["D"].width = 18
    evidence_sheet.column_dimensions["E"].width = 20
    evidence_sheet.column_dimensions["F"].width = 32
    evidence_sheet.column_dimensions["G"].width = 68
    evidence_sheet.column_dimensions["H"].width = 42
    evidence_sheet.column_dimensions["I"].width = 56

    image_streams: list[io.BytesIO] = []
    evidence_row = 2
    for result in public_results:
        for evidence in result.get("evidence_files") or []:
            evidence_sheet.append(
                [
                    _excel_text(result.get("host")),
                    result.get("port"),
                    _excel_text(result.get("protocol")),
                    _excel_text(result.get("service_name")),
                    _excel_text(evidence.get("type")),
                    _excel_text(evidence.get("file_name")),
                    _excel_text(evidence.get("sha256")),
                    _excel_text(evidence.get("source_url")),
                    "",
                ]
            )
            loaded = load_evidence(str(evidence.get("id") or ""))
            if loaded:
                try:
                    _, source_path = loaded
                    with PillowImage.open(source_path) as source_image:
                        source_image.load()
                        normalized = source_image.convert("RGB")
                        contained = ImageOps.contain(normalized, (400, 300))
                        canvas = PillowImage.new("RGB", (400, 300), "white")
                        left = (400 - contained.width) // 2
                        top = (300 - contained.height) // 2
                        canvas.paste(contained, (left, top))
                        image_stream = io.BytesIO()
                        canvas.save(image_stream, format="PNG")
                        image_stream.seek(0)
                    image_streams.append(image_stream)
                    excel_image = ExcelImage(image_stream)
                    excel_image.width = 400
                    excel_image.height = 300
                    evidence_sheet.add_image(excel_image, f"I{evidence_row}")
                    evidence_sheet.row_dimensions[evidence_row].height = 225
                except Exception as exc:  # noqa: BLE001 - keep the workbook usable if one image is corrupt.
                    evidence_sheet.cell(evidence_row, 9, f"Image unavailable: {str(exc)[:160]}")
            evidence_row += 1

    if evidence_row == 2:
        evidence_sheet.cell(2, 1, "No image evidence")
    evidence_sheet.auto_filter.ref = f"A1:I{max(1, evidence_row - 1)}"
    for row in evidence_sheet.iter_rows(min_row=2):
        for cell in row[:8]:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    output = io.BytesIO()
    workbook.save(output)
    return output.getvalue()


def format_results_ndjson(job: dict[str, Any], results: list[dict[str, Any]]) -> str:
    lines = [json.dumps({"type": "job", "job": job})]
    lines.extend(
        json.dumps({"type": "result", "result": result})
        for result in public_result_dicts(results)
    )
    return "\n".join(lines) + "\n"


def _extension_for_mime_type(mime_type: str) -> str:
    return {
        "image/png": ".png",
        "image/jpeg": ".jpg",
        "image/gif": ".gif",
        "image/webp": ".webp",
    }.get(mime_type, ".bin")


def _style_excel_header(sheet: Any, font_type: Any, fill_type: Any, alignment_type: Any) -> None:
    for cell in sheet[1]:
        cell.font = font_type(bold=True, color="FFFFFF")
        cell.fill = fill_type("solid", fgColor="0F766E")
        cell.alignment = alignment_type(horizontal="center", vertical="center")
    sheet.row_dimensions[1].height = 24


def _excel_text(value: Any) -> str:
    text = "" if value is None else str(value)
    if text.startswith(("=", "+", "-", "@")):
        return f"'{text}"
    return text
