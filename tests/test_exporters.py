import json
import tempfile
import unittest
import zipfile
from io import BytesIO
from pathlib import Path

from netprobe.exporters import (
    format_results_csv,
    format_results_csv_bundle,
    format_results_json,
    format_results_ndjson,
    format_results_xlsx,
)


class ExporterTests(unittest.TestCase):
    def test_public_exports_omit_internal_measurements_without_mutating_input(self):
        job = {"id": "scan-1"}
        result = {
            "scan_id": "scan-1",
            "host": "127.0.0.1",
            "port": 80,
            "protocol": "tcp",
            "state": "open",
            "latency_ms": 1.2,
            "service_name": "http",
            "service_confidence": 0.98,
            "tags": [],
            "evidence_files": [
                {
                    "id": "evidence-1",
                    "type": "manual",
                    "file_name": "proof.png",
                    "download_url": "/v1/evidence/evidence-1/content",
                }
            ],
        }

        json_result = json.loads(format_results_json(job, [result]))["results"][0]
        csv_header = format_results_csv([result]).splitlines()[0]
        ndjson_result = json.loads(format_results_ndjson(job, [result]).splitlines()[1])["result"]

        for public_result in (json_result, ndjson_result):
            self.assertNotIn("latency_ms", public_result)
            self.assertNotIn("service_confidence", public_result)
        self.assertNotIn("latency_ms", csv_header)
        self.assertNotIn("service_confidence", csv_header)
        self.assertEqual(json_result["evidence_files"][0]["file_name"], "proof.png")
        self.assertEqual(ndjson_result["evidence_files"][0]["id"], "evidence-1")
        self.assertIn("evidence_files", csv_header)
        self.assertEqual(result["latency_ms"], 1.2)
        self.assertEqual(result["service_confidence"], 0.98)

    def test_csv_evidence_bundle_contains_csv_manifest_and_image_files(self):
        with tempfile.TemporaryDirectory() as tmp:
            image_path = Path(tmp) / "evidence.png"
            image_path.write_bytes(b"\x89PNG\r\n\x1a\nbundled")
            evidence = {
                "id": "evidence-1",
                "type": "manual",
                "file_name": "proof.png",
                "mime_type": "image/png",
                "download_url": "/v1/evidence/evidence-1/content",
            }
            result = {
                "scan_id": "scan-1",
                "host": "127.0.0.1",
                "port": 80,
                "protocol": "tcp",
                "state": "open",
                "evidence_files": [evidence],
            }

            bundle = format_results_csv_bundle(
                {"id": "scan-1"},
                [result],
                load_evidence=lambda _evidence_id: (evidence, image_path),
            )

            with zipfile.ZipFile(BytesIO(bundle)) as archive:
                self.assertEqual(
                    set(archive.namelist()),
                    {"results.csv", "manifest.json", "evidence/evidence-1.png"},
                )
                csv_text = archive.read("results.csv").decode("utf-8-sig")
                self.assertIn("evidence/evidence-1.png", csv_text)
                self.assertEqual(archive.read("evidence/evidence-1.png"), image_path.read_bytes())

    def test_xlsx_export_embeds_images_in_evidence_sheet(self):
        from openpyxl import load_workbook
        from PIL import Image

        with tempfile.TemporaryDirectory() as tmp:
            image_path = Path(tmp) / "evidence.png"
            Image.new("RGB", (1042, 706), "navy").save(image_path, format="PNG")
            evidence = {
                "id": "evidence-1",
                "type": "manual",
                "file_name": "proof.png",
                "mime_type": "image/png",
                "sha256": "abc123",
            }
            result = {
                "scan_id": "scan-1",
                "host": "127.0.0.1",
                "port": 80,
                "protocol": "tcp",
                "state": "open",
                "service_name": "http",
                "note": "=SUM(1,1)",
                "evidence_files": [evidence],
            }

            payload = format_results_xlsx(
                {"id": "scan-1"},
                [result],
                load_evidence=lambda _evidence_id: (evidence, image_path),
            )
            workbook = load_workbook(BytesIO(payload))

            self.assertEqual(workbook.sheetnames, ["Results", "Evidence"])
            self.assertEqual(workbook["Results"]["B2"].value, "127.0.0.1")
            self.assertEqual(workbook["Results"]["J2"].value, "'=SUM(1,1)")
            self.assertEqual(workbook["Evidence"]["F2"].value, "proof.png")
            self.assertEqual(len(workbook["Evidence"]._images), 1)
            self.assertEqual(workbook["Evidence"]._images[0].width, 400)
            self.assertEqual(workbook["Evidence"]._images[0].height, 300)


if __name__ == "__main__":
    unittest.main()
